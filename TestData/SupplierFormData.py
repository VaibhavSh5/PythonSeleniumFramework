import os

import openpyxl


class SupplierFormData:
    Dict = {}
    list_dict = []
    supplier_form_data = [{"CompanyName": "Vaibhav Supplier Company 1", "CompanyLegalName": "Vaibhav Supplier "
                                                                                              "Company Legal Name 1",
                           "Street1": "Atlanta Street 1", "Street2": "Atlanta Street 2", "City": "Atlanta City 1",
                           "Country": "United States of America", "State/Province": "Georgia", "Zip/PostalCode":
                               "30309", "Country/Region": "USA"},
                          {"CompanyName": "Vaibhav Supplier Company 2", "CompanyLegalName": "Vaibhav Supplier "
                                                                                              "Company Legal Name 2",
                           "Street1": "Atlanta Street 1.1", "Street2": "Atlanta Street 2.1", "City": "Atlanta City 2",
                           "Country": "India", "State/Province": "Uttar Pradesh", "Zip/PostalCode":
                               "201307", "Country/Region": "INDIA"}]

    @staticmethod
    def getsupplierdata():
        workbook = openpyxl.load_workbook(os.path.dirname(os.path.dirname(__file__))+"/ExcelDataDriven/ExcelData.xlsx")
        sheet = workbook.active

        for i in range(1, sheet.max_row+1):
            # if sheet.cell(row=i, column=1).value != "CompanyName":
            for j in range(1, sheet.max_column+1):
                SupplierFormData.Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            SupplierFormData.list_dict.append(SupplierFormData.Dict)

        return SupplierFormData.list_dict
